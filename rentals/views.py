from django.http import HttpResponse
from django.template.loader import get_template
from django.utils.timezone import now
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import calendar
import os
from django.conf import settings
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.forms import UserCreationForm
from django.contrib import messages
from django.db.models import Sum, Q
from django.utils.translation import gettext as _
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta
from .models import MainCategory, SubCategory, Unit, Contract, Payment, Tenant, Expense
from .forms import MainCategoryForm, SubCategoryForm, UnitForm, TenantForm, ContractForm, PaymentForm, ExpenseForm
from django.conf import settings
import arabic_reshaper
from bidi.algorithm import get_display
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from django.db.models import Q

# ============================================================
# دوال مساعدة للضريبة الربعية
# ============================================================

def get_quarter_dates(year, quarter):
    if quarter == 1:
        start_date = datetime(year, 1, 1)
        end_date = datetime(year, 3, 31)
    elif quarter == 2:
        start_date = datetime(year, 4, 1)
        end_date = datetime(year, 6, 30)
    elif quarter == 3:
        start_date = datetime(year, 7, 1)
        end_date = datetime(year, 9, 30)
    elif quarter == 4:
        start_date = datetime(year, 10, 1)
        end_date = datetime(year, 12, 31)
    else:
        return None, None
    return start_date.date(), end_date.date()

def get_quarter_tax(year, quarter):
    start_date, end_date = get_quarter_dates(year, quarter)
    if not start_date:
        return {'due': 0, 'collected': 0}
    
    contracts = Contract.objects.filter(
        start_date__lte=end_date,
        is_active=True
    ).exclude(start_date__gt=end_date)
    
    tax_due = 0
    for contract in contracts:
        if contract.start_date <= end_date and contract.end_date >= start_date:
            overlap_start = max(contract.start_date, start_date)
            overlap_end = min(contract.end_date, end_date)
            months = (overlap_end.year - overlap_start.year) * 12 + (overlap_end.month - overlap_start.month) + 1
            tax_due += contract.tax_amount_monthly * months
    
    payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date
    )
    tax_collected = 0
    for payment in payments:
        contract = payment.contract
        if contract.has_tax:
            tax_collected += payment.amount_paid * (contract.tax_rate / 100) / (1 + contract.tax_rate/100)
    
    return {'due': round(tax_due, 2), 'collected': round(tax_collected, 2)}

# ============================================================
# دوال المصادقة (لا تحتاج login_required)
# ============================================================

def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect('home')
        else:
            messages.error(request, _('اسم المستخدم أو كلمة المرور غير صحيحة'))
    return render(request, 'rentals/login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

def register_view(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'تم إنشاء حساب لـ {username}')
            return redirect('login')
    else:
        form = UserCreationForm()
    return render(request, 'rentals/register.html', {'form': form})

# ============================================================
# دوال محمية (يجب تسجيل الدخول)
# ============================================================

@login_required
def home(request):
    current_year = datetime.now().year
    
    total_categories = MainCategory.objects.count()
    total_units = Unit.objects.count()
    rented_units = Unit.objects.filter(is_rented=True).count()
    
    contracts = Contract.objects.filter(is_active=True)
    total_expected = sum(c.total_expected for c in contracts)
    total_paid = Payment.objects.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    total_remaining = total_expected - total_paid
    
    total_tax_due = 0
    total_tax_collected = 0
    quarters = []
    for q in range(1, 5):
        tax_data = get_quarter_tax(current_year, q)
        quarters.append({
            'quarter': q,
            'due': tax_data['due'],
            'collected': tax_data['collected'],
        })
        total_tax_due += tax_data['due']
        total_tax_collected += tax_data['collected']
    
    subcategories = SubCategory.objects.filter(is_active=True).prefetch_related('units', 'expenses')
    subcategories_data = []
    for sub in subcategories:
        units_count = sub.units.count()
        rented_count = sub.units.filter(is_rented=True).count()
        contracts_sub = Contract.objects.filter(unit__sub_category=sub, is_active=True)
        total_rent_expected = sum(c.total_expected for c in contracts_sub)
        total_expense = sub.total_expenses()
        refundable_tax = sub.total_refundable_tax()
        subcategories_data.append({
            'id': sub.id,
            'name': sub.name,
            'main_category': sub.main_category.name,
            'units_count': units_count,
            'rented_count': rented_count,
            'total_rent_expected': total_rent_expected,
            'total_expense': total_expense,
            'refundable_tax': refundable_tax,
        })
    
    context = {
        'total_categories': total_categories,
        'total_units': total_units,
        'rented_units': rented_units,
        'total_expected': total_expected,
        'total_paid': total_paid,
        'total_remaining': total_remaining,
        'total_tax_due': total_tax_due,
        'total_tax_collected': total_tax_collected,
        'quarters': quarters,
        'current_year': current_year,
        'subcategories': subcategories_data,
    }
    return render(request, 'rentals/home.html', context)

@login_required
def main_categories(request):
    categories = MainCategory.objects.all()
    return render(request, 'rentals/main_categories.html', {'categories': categories})

@login_required
def main_category_detail(request, pk):
    category = get_object_or_404(MainCategory, pk=pk)
    subcategories = category.subcategories.all()
    
    units_count = Unit.objects.filter(sub_category__main_category=category).count()
    rented_count = Unit.objects.filter(sub_category__main_category=category, is_rented=True).count()
    
    contracts = Contract.objects.filter(
        unit__sub_category__main_category=category,
        is_active=True
    )
    total_expected = sum(c.total_expected for c in contracts)
    
    context = {
        'category': category,
        'subcategories': subcategories,
        'units_count': units_count,
        'rented_count': rented_count,
        'total_expected': total_expected,
    }
    return render(request, 'rentals/main_category_detail.html', context)

@login_required
def subcategory_detail(request, pk):
    subcategory = get_object_or_404(SubCategory, pk=pk)
    units = subcategory.units.all()
    rented_units = units.filter(is_rented=True).count()
    
    contracts = Contract.objects.filter(unit__sub_category=subcategory, is_active=True)
    total_expected = sum(c.total_expected for c in contracts)
    total_paid = Payment.objects.filter(contract__unit__sub_category=subcategory).aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    
    context = {
        'subcategory': subcategory,
        'units': units,
        'rented_units': rented_units,
        'total_expected': total_expected,
        'total_paid': total_paid,
    }
    return render(request, 'rentals/subcategory_detail.html', context)

@login_required
def unit_detail(request, pk):
    unit = get_object_or_404(Unit, pk=pk)
    current_contract = unit.current_contract
    payments = []
    if current_contract:
        payments = current_contract.payments.all().order_by('-payment_date')
    
    context = {
        'unit': unit,
        'current_contract': current_contract,
        'payments': payments,
    }
    return render(request, 'rentals/unit_detail.html', context)

@login_required
def add_unit_flow(request):
    categories = MainCategory.objects.all()
    return render(request, 'rentals/add_unit_flow.html', {'categories': categories})

@login_required
@permission_required('rentals.add_maincategory', raise_exception=True)
def add_main_category(request):
    if request.method == 'POST':
        form = MainCategoryForm(request.POST)
        if form.is_valid():
            category = form.save()
            messages.success(request, _('تم إضافة الفئة الرئيسية بنجاح.'))
            return redirect('choose_subcategory', category_id=category.id)
    else:
        form = MainCategoryForm()
    return render(request, 'rentals/add_main_category.html', {'form': form})

@login_required
def choose_subcategory(request, category_id):
    category = get_object_or_404(MainCategory, pk=category_id)
    subcategories = category.subcategories.all()
    if request.method == 'POST':
        subcategory_id = request.POST.get('subcategory_id')
        if subcategory_id:
            subcategory = get_object_or_404(SubCategory, pk=subcategory_id)
            return redirect('add_unit', subcategory_id=subcategory.id)
        else:
            return redirect('add_sub_category', category_id=category.id)
    context = {'category': category, 'subcategories': subcategories}
    return render(request, 'rentals/choose_subcategory.html', context)

@login_required
@permission_required('rentals.add_subcategory', raise_exception=True)
def add_sub_category(request, category_id):
    category = get_object_or_404(MainCategory, pk=category_id)
    if request.method == 'POST':
        form = SubCategoryForm(request.POST)
        if form.is_valid():
            sub = form.save(commit=False)
            sub.main_category = category
            sub.save()
            messages.success(request, _('تم إضافة القسم الداخلي بنجاح.'))
            return redirect('add_unit', subcategory_id=sub.id)
    else:
        form = SubCategoryForm()
    return render(request, 'rentals/add_sub_category.html', {'form': form, 'category': category})

@login_required
@permission_required('rentals.add_unit', raise_exception=True)
def add_unit(request, subcategory_id):
    subcategory = get_object_or_404(SubCategory, pk=subcategory_id)
    if request.method == 'POST':
        form = UnitForm(request.POST, request.FILES)
        if form.is_valid():
            unit = form.save(commit=False)
            unit.sub_category = subcategory
            unit.save()
            messages.success(request, _('تم إضافة الوحدة بنجاح.'))
            request.session['new_unit_id'] = unit.id
            return redirect('add_tenant')
    else:
        form = UnitForm()
    return render(request, 'rentals/add_unit.html', {'form': form, 'subcategory': subcategory})

@login_required
@permission_required('rentals.add_tenant', raise_exception=True)
def add_tenant(request):
    if request.method == 'POST':
        form = TenantForm(request.POST, request.FILES)
        if form.is_valid():
            identity_number = normalize_arabic_numbers(form.cleaned_data['identity_number'])
            
            if Tenant.objects.filter(identity_number=identity_number, is_deleted=False).exists():
                messages.error(request, _('رقم الهوية مستخدم مسبقاً من قبل مستأجر نشط.'))
                return render(request, 'rentals/add_tenant.html', {'form': form})
            else:
                form.instance.identity_number = identity_number
                tenant = form.save()
                request.session['new_tenant_id'] = tenant.id
                return redirect('add_contract')
    else:
        form = TenantForm()
    return render(request, 'rentals/add_tenant.html', {'form': form})

@login_required
@permission_required('rentals.add_contract', raise_exception=True)
def add_contract_for_unit(request, unit_id):
    """بدء إضافة عقد لوحدة موجودة (غير مؤجرة)"""
    unit = get_object_or_404(Unit, pk=unit_id)
    
    # التحقق من أن الوحدة غير مؤجرة
    if unit.is_rented:
        messages.warning(request, _('الوحدة مؤجرة حالياً. يجب إنهاء العقد الحالي أولاً.'))
        return redirect('unit_detail', pk=unit.id)
    
    # تخزين معرف الوحدة في الجلسة للاستخدام في الخطوات التالية
    request.session['new_unit_id'] = unit.id
    
    # التوجيه إلى إضافة مستأجر (أو يمكن تخطي المستأجر إذا أردنا اختيار مستأجر موجود)
    return redirect('add_tenant')

@login_required
@permission_required('rentals.add_contract', raise_exception=True)
def add_contract(request):
    unit_id = request.session.get('new_unit_id')
    tenant_id = request.session.get('new_tenant_id')
    if not unit_id or not tenant_id:
        messages.error(request, _('الرجاء البدء من البداية.'))
        return redirect('home')
    unit = get_object_or_404(Unit, pk=unit_id)
    tenant = get_object_or_404(Tenant, pk=tenant_id)
    if request.method == 'POST':
        form = ContractForm(request.POST, request.FILES)
        if form.is_valid():
            contract = form.save(commit=False)
            contract.unit = unit
            contract.tenant = tenant
            contract.save()
            unit.is_rented = True
            unit.save()
            request.session['new_contract_id'] = contract.id
            messages.success(request, _('تم إضافة العقد بنجاح.'))
            return redirect('add_payment')
    else:
        form = ContractForm()
    return render(request, 'rentals/add_contract.html', {'form': form, 'unit': unit, 'tenant': tenant})

@login_required
@permission_required('rentals.add_payment', raise_exception=True)
def add_payment(request):
    contract_id = request.GET.get('contract') or request.session.get('new_contract_id')
    if not contract_id:
        messages.error(request, _('لم يتم تحديد العقد.'))
        return redirect('home')
    contract = get_object_or_404(Contract, pk=contract_id)
    if request.method == 'POST':
        if 'skip' in request.POST:
            for key in ['new_unit_id', 'new_tenant_id', 'new_contract_id']:
                request.session.pop(key, None)
            messages.success(request, _('تم إتمام عملية الإضافة بنجاح.'))
            return redirect('unit_detail', pk=contract.unit.id)
        form = PaymentForm(request.POST, request.FILES)
        if form.is_valid():
            payment = form.save(commit=False)
            payment.contract = contract
            payment.save()
            messages.success(request, _('تم إضافة الدفعة بنجاح.'))
        for key in ['new_unit_id', 'new_tenant_id', 'new_contract_id']:
            request.session.pop(key, None)
        return redirect('unit_detail', pk=contract.unit.id)
    else:
        initial_data = {
            'payment_date': date.today(),
            'for_period_start': contract.start_date,
            'for_period_end': contract.start_date + relativedelta(months=1) - timedelta(days=1),
        }
        form = PaymentForm(initial=initial_data)
    return render(request, 'rentals/add_payment.html', {'form': form, 'contract': contract})

@login_required
@permission_required('rentals.add_expense', raise_exception=True)
def add_expense(request, subcategory_id):
    subcategory = get_object_or_404(SubCategory, pk=subcategory_id)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES)
        if form.is_valid():
            expense = form.save(commit=False)
            expense.sub_category = subcategory
            expense.save()
            messages.success(request, _('تم إضافة المصروف بنجاح.'))
            return redirect('subcategory_detail', pk=subcategory.id)
    else:
        form = ExpenseForm(initial={'date': date.today()})
    return render(request, 'rentals/add_expense.html', {'form': form, 'subcategory': subcategory})

@login_required
@permission_required('rentals.view_expense', raise_exception=False)  # للعرض فقط
def profit_loss_report(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    try:
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        else:
            start_date = date.today().replace(day=1)
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        else:
            end_date = date.today()
    except:
        start_date = date.today().replace(day=1)
        end_date = date.today()
    payments = Payment.objects.filter(payment_date__gte=start_date, payment_date__lte=end_date)
    total_rent = payments.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    expenses = Expense.objects.filter(date__gte=start_date, date__lte=end_date)
    total_expense = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    refundable_tax = 0
    for exp in expenses.filter(tax_refundable=True):
        if exp.has_tax:
            tax = exp.amount * (exp.tax_rate / 100) / (1 + exp.tax_rate/100)
            refundable_tax += tax
    net_profit = total_rent - total_expense + refundable_tax
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'total_rent': total_rent,
        'total_expense': total_expense,
        'refundable_tax': refundable_tax,
        'net_profit': net_profit,
        'payments': payments.order_by('-payment_date')[:50],
        'expenses': expenses.order_by('-date')[:50],
    }
    return render(request, 'rentals/profit_loss_report.html', context)

# ============================================================
# دوال التعديل (Edit)
# ============================================================

@login_required
@permission_required('rentals.change_maincategory', raise_exception=True)
def edit_main_category(request, pk):
    category = get_object_or_404(MainCategory, pk=pk)
    if request.method == 'POST':
        form = MainCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, _('تم تعديل الفئة الرئيسية بنجاح.'))
            return redirect('main_category_detail', pk=category.id)
    else:
        form = MainCategoryForm(instance=category)
    return render(request, 'rentals/edit_main_category.html', {'form': form, 'category': category})

@login_required
@permission_required('rentals.change_subcategory', raise_exception=True)
def edit_sub_category(request, pk):
    sub = get_object_or_404(SubCategory, pk=pk)
    if request.method == 'POST':
        form = SubCategoryForm(request.POST, instance=sub)
        if form.is_valid():
            form.save()
            messages.success(request, _('تم تعديل القسم الداخلي بنجاح.'))
            return redirect('subcategory_detail', pk=sub.id)
    else:
        form = SubCategoryForm(instance=sub)
    return render(request, 'rentals/edit_sub_category.html', {'form': form, 'subcategory': sub})

@login_required
@permission_required('rentals.change_unit', raise_exception=True)
def edit_unit(request, pk):
    unit = get_object_or_404(Unit, pk=pk)
    if request.method == 'POST':
        form = UnitForm(request.POST, request.FILES, instance=unit)
        if form.is_valid():
            form.save()
            messages.success(request, _('تم تعديل الوحدة بنجاح.'))
            return redirect('unit_detail', pk=unit.id)
    else:
        form = UnitForm(instance=unit)
    return render(request, 'rentals/edit_unit.html', {'form': form, 'unit': unit})

from django.db import IntegrityError
import unicodedata

def normalize_arabic_numbers(text):
    """تحويل الأرقام العربية (٠-٩) إلى أرقام إنجليزية (0-9)"""
    if text:
        text = str(text).strip()
        # تحويل الأرقام العربية إلى إنجليزية
        arabic_numbers = str.maketrans('٠١٢٣٤٥٦٧٨٩', '0123456789')
        text = text.translate(arabic_numbers)
        # تحويل الأرقام الفارسية/العربية الأخرى
        persian_numbers = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
        text = text.translate(persian_numbers)
    return text

@login_required
@permission_required('rentals.change_tenant', raise_exception=True)
def edit_tenant(request, pk):
    tenant = get_object_or_404(Tenant, pk=pk, is_deleted=False)
    
    if request.method == 'POST':
        form = TenantForm(request.POST, request.FILES, instance=tenant)
        if form.is_valid():
            # تطبيع رقم الهوية
            identity_number = normalize_arabic_numbers(form.cleaned_data['identity_number'])
            
            # التحقق من عدم وجود مستأجر آخر نشط بنفس الرقم
            existing = Tenant.objects.filter(
                identity_number=identity_number, 
                is_deleted=False
            ).exclude(pk=pk)
            
            if existing.exists():
                messages.error(request, _('رقم الهوية مستخدم مسبقاً من قبل مستأجر آخر.'))
                return render(request, 'rentals/edit_tenant.html', {'form': form, 'tenant': tenant})
            
            # تعيين الرقم الطبيعي للنموذج
            form.instance.identity_number = identity_number
            
            try:
                form.save()
                messages.success(request, _('تم تعديل بيانات المستأجر بنجاح.'))
                contract = tenant.contracts.first()
                if contract:
                    return redirect('unit_detail', pk=contract.unit.id)
                return redirect('home')
            except IntegrityError as e:
                messages.error(request, _('حدث خطأ في حفظ البيانات. الرقم مستخدم بالفعل أو مشكلة في القيد.'))
                return render(request, 'rentals/edit_tenant.html', {'form': form, 'tenant': tenant})
    else:
        form = TenantForm(instance=tenant)
    
    return render(request, 'rentals/edit_tenant.html', {'form': form, 'tenant': tenant})
@login_required
@permission_required('rentals.change_contract', raise_exception=True)
def edit_contract(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    if request.method == 'POST':
        form = ContractForm(request.POST, request.FILES, instance=contract)
        if form.is_valid():
            form.save()
            messages.success(request, _('تم تعديل العقد بنجاح.'))
            return redirect('unit_detail', pk=contract.unit.id)
    else:
        form = ContractForm(instance=contract)
    return render(request, 'rentals/edit_contract.html', {'form': form, 'contract': contract})

@login_required
@permission_required('rentals.change_payment', raise_exception=True)
def edit_payment(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    if request.method == 'POST':
        form = PaymentForm(request.POST, request.FILES, instance=payment)
        if form.is_valid():
            form.save()
            messages.success(request, _('تم تعديل الدفعة بنجاح.'))
            return redirect('unit_detail', pk=payment.contract.unit.id)
    else:
        form = PaymentForm(instance=payment)
    return render(request, 'rentals/edit_payment.html', {'form': form, 'payment': payment})

@login_required
@permission_required('rentals.change_expense', raise_exception=True)
def edit_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    if request.method == 'POST':
        form = ExpenseForm(request.POST, request.FILES, instance=expense)
        if form.is_valid():
            form.save()
            messages.success(request, _('تم تعديل المصروف بنجاح.'))
            return redirect('subcategory_detail', pk=expense.sub_category.id)
    else:
        form = ExpenseForm(instance=expense)
    return render(request, 'rentals/edit_expense.html', {'form': form, 'expense': expense})

# ============================================================
# دوال الحذف (Delete)
# ============================================================

@login_required
@permission_required('rentals.delete_maincategory', raise_exception=True)
def delete_main_category(request, pk):
    category = get_object_or_404(MainCategory, pk=pk)
    if request.method == 'POST':
        category.delete()
        messages.success(request, _('تم حذف الفئة الرئيسية بنجاح.'))
        return redirect('main_categories')
    return render(request, 'rentals/delete_confirm.html', {'object': category, 'type': 'main_category'})

@login_required
@permission_required('rentals.delete_subcategory', raise_exception=True)
def delete_sub_category(request, pk):
    sub = get_object_or_404(SubCategory, pk=pk)
    main_cat_id = sub.main_category.id
    if request.method == 'POST':
        sub.delete()
        messages.success(request, _('تم حذف القسم الداخلي بنجاح.'))
        return redirect('main_category_detail', pk=main_cat_id)
    return render(request, 'rentals/delete_confirm.html', {'object': sub, 'type': 'sub_category'})

@login_required
@permission_required('rentals.delete_unit', raise_exception=True)
def delete_unit(request, pk):
    unit = get_object_or_404(Unit, pk=pk)
    sub_id = unit.sub_category.id
    if request.method == 'POST':
        unit.delete()
        messages.success(request, _('تم حذف الوحدة بنجاح.'))
        return redirect('subcategory_detail', pk=sub_id)
    return render(request, 'rentals/delete_confirm.html', {'object': unit, 'type': 'unit'})

@login_required
@permission_required('rentals.delete_tenant', raise_exception=True)
def delete_tenant(request, pk):
    tenant = get_object_or_404(Tenant, pk=pk)
    if request.method == 'POST':
        # بدلاً من الحذف الفعلي، نضع علامة كمحذوف
        tenant.is_deleted = True
        tenant.save()
        messages.success(request, _('تم حذف المستأجر بنجاح.'))
        return redirect('home')
    return render(request, 'rentals/delete_confirm.html', {'object': tenant, 'type': 'tenant'})
@login_required
@permission_required('rentals.delete_contract', raise_exception=True)
def delete_contract(request, pk):
    contract = get_object_or_404(Contract, pk=pk)
    unit_id = contract.unit.id
    if request.method == 'POST':
        contract.delete()
        unit = Unit.objects.get(pk=unit_id)
        unit.is_rented = False
        unit.save()
        messages.success(request, _('تم حذف العقد بنجاح.'))
        return redirect('unit_detail', pk=unit_id)
    return render(request, 'rentals/delete_confirm.html', {'object': contract, 'type': 'contract'})

@login_required
@permission_required('rentals.delete_payment', raise_exception=True)
def delete_payment(request, pk):
    payment = get_object_or_404(Payment, pk=pk)
    unit_id = payment.contract.unit.id
    if request.method == 'POST':
        payment.delete()
        messages.success(request, _('تم حذف الدفعة بنجاح.'))
        return redirect('unit_detail', pk=unit_id)
    return render(request, 'rentals/delete_confirm.html', {'object': payment, 'type': 'payment'})

@login_required
@permission_required('rentals.delete_expense', raise_exception=True)
def delete_expense(request, pk):
    expense = get_object_or_404(Expense, pk=pk)
    sub_id = expense.sub_category.id
    if request.method == 'POST':
        expense.delete()
        messages.success(request, _('تم حذف المصروف بنجاح.'))
        return redirect('subcategory_detail', pk=sub_id)
    return render(request, 'rentals/delete_confirm.html', {'object': expense, 'type': 'expense'})



# ============================================================
# دوال مساعدة للتقارير
# ============================================================

def prepare_arabic_text(text):
    """تجهيز النص العربي للعرض في PDF"""
    if text:
        reshaped_text = arabic_reshaper.reshape(str(text))
        return get_display(reshaped_text)
    return ""

def get_month_name(month_number):
    """إرجاع اسم الشهر بالعربية"""
    months = {
        1: 'يناير', 2: 'فبراير', 3: 'مارس', 4: 'إبريل',
        5: 'مايو', 6: 'يونيو', 7: 'يوليو', 8: 'أغسطس',
        9: 'سبتمبر', 10: 'أكتوبر', 11: 'نوفمبر', 12: 'ديسمبر'
    }
    return months.get(month_number, '')
# ============================================================
# صفحة التقارير الرئيسية
# ============================================================

@login_required
def reports_dashboard(request):
    """الصفحة الرئيسية للتقارير"""
    current_year = datetime.now().year
    years = range(current_year - 5, current_year + 1)
    months = range(1, 13)
    
    context = {
        'years': years,
        'months': months,
        'current_year': current_year,
        'current_month': datetime.now().month,
    }
    return render(request, 'rentals/reports/dashboard.html', context)

# ============================================================
# تقرير الإيجارات الشهري
# ============================================================

@login_required
def rent_report(request):
    """تقرير الإيجارات الشهري"""
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year+1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month+1, 1) - timedelta(days=1)
    
    payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date
    ).select_related('contract__unit', 'contract__tenant')
    
    total_payments = payments.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0
    
    # حساب العقود النشطة يدوياً بدلاً من استخدام end_date في الفلتر
    all_contracts = Contract.objects.filter(is_active=True)
    active_contracts = []
    for contract in all_contracts:
        # حساب نهاية العقد يدوياً
        contract_end = contract.start_date + relativedelta(months=contract.lease_duration_months) - timedelta(days=1)
        if contract.start_date <= end_date and contract_end >= start_date:
            active_contracts.append(contract)
    
    expected_rent = 0
    for contract in active_contracts:
        contract_end = contract.start_date + relativedelta(months=contract.lease_duration_months) - timedelta(days=1)
        overlap_start = max(contract.start_date, start_date)
        overlap_end = min(contract_end, end_date)
        if overlap_start <= overlap_end:
            expected_rent += contract.total_monthly_with_tax
    
    current_year = datetime.now().year
    years = range(current_year - 5, current_year + 1)
    months = range(1, 13)
    
    context = {
        'year': year,
        'month': month,
        'month_name': get_month_name(month),
        'payments': payments,
        'total_payments': total_payments,
        'expected_rent': expected_rent,
        'start_date': start_date,
        'end_date': end_date,
        'years': years,
        'months': months,
    }
    return render(request, 'rentals/reports/rent_report.html', context)

# ============================================================
# تقرير المصاريف الشهري
# ============================================================

@login_required
def expense_report(request):
    """تقرير المصاريف الشهري"""
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year+1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month+1, 1) - timedelta(days=1)
    
    expenses = Expense.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).select_related('sub_category')
    
    total_expense = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    total_refundable_tax = 0
    for exp in expenses.filter(tax_refundable=True):
        total_refundable_tax += exp.tax_amount
    
    current_year = datetime.now().year
    years = range(current_year - 5, current_year + 1)
    months = range(1, 13)
    
    context = {
        'year': year,
        'month': month,
        'month_name': get_month_name(month),
        'expenses': expenses,
        'total_expense': total_expense,
        'total_refundable_tax': total_refundable_tax,
        'start_date': start_date,
        'end_date': end_date,
        'years': years,
        'months': months,
    }
    return render(request, 'rentals/reports/expense_report.html', context)

# ============================================================
# تقرير الضريبة
# ============================================================

@login_required
def tax_report(request):
    """تقرير الضريبة (شهري / ربع سنوي) - مع حساب صافي الضريبة"""
    report_type = request.GET.get('type', 'monthly')
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    quarter = int(request.GET.get('quarter', 1))
    
    # تحديد تاريخ البداية والنهاية
    if report_type == 'monthly':
        start_date = date(year, month, 1)
        if month == 12:
            end_date = date(year+1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(year, month+1, 1) - timedelta(days=1)
        period_name = f"{get_month_name(month)} {year}"
    else:  # quarterly
        if quarter == 1:
            start_date = date(year, 1, 1)
            end_date = date(year, 3, 31)
        elif quarter == 2:
            start_date = date(year, 4, 1)
            end_date = date(year, 6, 30)
        elif quarter == 3:
            start_date = date(year, 7, 1)
            end_date = date(year, 9, 30)
        else:
            start_date = date(year, 10, 1)
            end_date = date(year, 12, 31)
        period_name = f"الربع {quarter} {year}"
    
    # ------------------------------------------------------------
    # 1. الضريبة المستحقة (حسب تواريخ بداية دورات السداد)
    # ------------------------------------------------------------
    all_contracts = Contract.objects.filter(is_active=True)
    tax_due = 0
    tax_due_details = []
    
    for contract in all_contracts:
        # مدة دورة السداد بالأشهر
        interval_map = {'monthly': 1, 'quarterly': 3, 'half_yearly': 6, 'yearly': 12}
        interval_months = interval_map.get(contract.payment_interval, 1)
        
        contract_end = contract.start_date + relativedelta(months=contract.lease_duration_months) - timedelta(days=1)
        
        # إذا العقد لا يغطي الفترة، نتخطاه
        if contract.start_date > end_date or contract_end < start_date:
            continue
        
        # توليد تواريخ الاستحقاق
        due_date = contract.start_date
        while due_date <= contract_end:
            if start_date <= due_date <= end_date:
                tax_for_period = contract.tax_amount_monthly * interval_months
                tax_due += tax_for_period
                tax_due_details.append({
                    'contract': contract,
                    'date': due_date,
                    'amount': tax_for_period,
                })
            due_date += relativedelta(months=interval_months)
    
    # ------------------------------------------------------------
    # 2. الضريبة المحصلة (من الدفعات المسجلة)
    # ------------------------------------------------------------
    payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date
    )
    tax_collected = 0
    for payment in payments:
        if payment.contract.has_tax:
            tax_collected += payment.amount_paid * (payment.contract.tax_rate / 100) / (1 + payment.contract.tax_rate/100)
    
    # ------------------------------------------------------------
    # 3. الضريبة المستردة (من المصاريف)
    # ------------------------------------------------------------
    expenses = Expense.objects.filter(
        date__gte=start_date,
        date__lte=end_date,
        tax_refundable=True
    )
    tax_refunded = sum(exp.tax_amount for exp in expenses)
    
    # ------------------------------------------------------------
    # 4. صافي الضريبة المستحق للحكومة
    # ------------------------------------------------------------
    net_tax = tax_due - tax_refunded   # هذا هو المهم
    
    # ------------------------------------------------------------
    # إعداد متغيرات للقالب
    # ------------------------------------------------------------
    current_year = datetime.now().year
    years = range(current_year - 5, current_year + 1)
    months = range(1, 13)
    
    context = {
        'report_type': report_type,
        'year': year,
        'month': month,
        'quarter': quarter,
        'period_name': period_name,
        'start_date': start_date,
        'end_date': end_date,
        'tax_due': tax_due,
        'tax_collected': tax_collected,
        'tax_refunded': tax_refunded,
        'net_tax': net_tax,                 # تم التأكيد على هذا المتغير
        'tax_due_details': tax_due_details, # اختياري
        'years': years,
        'months': months,
    }
    return render(request, 'rentals/reports/tax_report.html', context)

# ============================================================
# تصدير إلى Excel
# ============================================================

@login_required
def export_rent_excel(request):
    """تصدير تقرير الإيجارات إلى Excel"""
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year+1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month+1, 1) - timedelta(days=1)
    
    payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date
    ).select_related('contract__unit', 'contract__tenant')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"تقرير إيجارات {get_month_name(month)} {year}"
    
    headers = ['التاريخ', 'رقم العقد', 'الوحدة', 'المستأجر', 'المبلغ', 'طريقة الدفع', 'ملاحظات']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, payment in enumerate(payments, 2):
        ws.cell(row=row_num, column=1).value = payment.payment_date
        ws.cell(row=row_num, column=2).value = payment.contract.contract_number
        ws.cell(row=row_num, column=3).value = payment.contract.unit.unit_number
        ws.cell(row=row_num, column=4).value = payment.contract.tenant.name
        ws.cell(row=row_num, column=5).value = float(payment.amount_paid)
        ws.cell(row=row_num, column=6).value = payment.get_payment_method_display()
        ws.cell(row=row_num, column=7).value = payment.notes
    
    for col_num in range(1, 8):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="rent_report_{year}_{month}.xlsx"'
    wb.save(response)
    return response

@login_required
def export_expense_excel(request):
    """تصدير تقرير المصاريف إلى Excel"""
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year+1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month+1, 1) - timedelta(days=1)
    
    expenses = Expense.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).select_related('sub_category')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"تقرير مصاريف {get_month_name(month)} {year}"
    
    headers = ['التاريخ', 'القسم', 'البيان', 'المبلغ', 'ضريبة مستردة', 'ملاحظات']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for row_num, exp in enumerate(expenses, 2):
        ws.cell(row=row_num, column=1).value = exp.date
        ws.cell(row=row_num, column=2).value = exp.sub_category.name
        ws.cell(row=row_num, column=3).value = exp.description
        ws.cell(row=row_num, column=4).value = float(exp.amount)
        ws.cell(row=row_num, column=5).value = float(exp.tax_amount) if exp.tax_refundable else 0
        ws.cell(row=row_num, column=6).value = exp.notes
    
    for col_num in range(1, 7):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="expense_report_{year}_{month}.xlsx"'
    wb.save(response)
    return response

# ============================================================
# تصدير إلى PDF
# ============================================================

@login_required
def export_rent_pdf(request):
    """تصدير تقرير الإيجارات إلى PDF"""
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year+1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month+1, 1) - timedelta(days=1)
    
    payments = Payment.objects.filter(
        payment_date__gte=start_date,
        payment_date__lte=end_date
    ).select_related('contract__unit', 'contract__tenant')
    
    total_payments = payments.aggregate(Sum('amount_paid'))['amount_paid__sum'] or 0

    # ------------------- البحث عن خط عربي -------------------
    possible_paths = [
        '/Library/Fonts/Arial.ttf',                          # macOS
        '/System/Library/Fonts/Supplemental/Arial.ttf',      # macOS
        'C:/Windows/Fonts/Arial.ttf',                        # Windows
        '/usr/share/fonts/truetype/msttcorefonts/Arial.ttf', # Linux
    ]
    font_registered = False
    for path in possible_paths:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont('Arabic', path))
            font_registered = True
            break
    if not font_registered:
        # في حال لم يتم العثور على خط، نستخدم Helvetica (قد لا يدعم العربية)
        pdfmetrics.registerFont(TTFont('Arabic', 'Helvetica'))
    # ---------------------------------------------------------

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="rent_report_{year}_{month}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
    elements = []
    
    styles = getSampleStyleSheet()
    styles['Title'].fontName = 'Arabic'
    title_style = styles['Title']
    title_style.alignment = 1
    
    # عنوان التقرير (مع reshape)
    title_text = f"تقرير الإيجارات - {get_month_name(month)} {year}"
    title = Paragraph(prepare_arabic_text(title_text), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # معلومات الفترة والإجمالي (مع reshape)
    info_style = ParagraphStyle('info', parent=styles['Normal'], fontName='Arabic', alignment=1)
    info_text = f"الفترة: {start_date} إلى {end_date} | الإجمالي: {total_payments:,.2f}"
    info = Paragraph(prepare_arabic_text(info_text), info_style)
    elements.append(info)
    elements.append(Spacer(1, 0.5*cm))
    
    # بيانات الجدول
    data = []
    headers = ['التاريخ', 'رقم العقد', 'الوحدة', 'المستأجر', 'المبلغ', 'طريقة الدفع']
    reshaped_headers = [prepare_arabic_text(h) for h in headers]
    data.append(reshaped_headers)
    
    for p in payments:
        row = [
            str(p.payment_date),
            p.contract.contract_number,
            p.contract.unit.unit_number,
            prepare_arabic_text(p.contract.tenant.name),          # اسم المستأجر
            f"{p.amount_paid:,.2f}",
            prepare_arabic_text(p.get_payment_method_display()), # طريقة الدفع
        ]
        data.append(row)
    
    table = Table(data, colWidths=[3*cm, 3*cm, 3*cm, 4*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

from django.db.models import Q  # تأكد من وجود هذا الاستيراد مع بقية الاستيرادات

from django.db.models import Q

@login_required
def advanced_search(request):
    """
    صفحة البحث المتقدم مع التصفية حسب معايير متعددة.
    """
    query = request.GET.get('q', '')
    entity = request.GET.get('entity', 'all')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    status = request.GET.get('status', '')
    unit_type = request.GET.get('unit_type', '')
    main_category = request.GET.get('main_category', '')
    sub_category = request.GET.get('sub_category', '')

    results = {
        'contracts': [],
        'units': [],
        'tenants': [],
        'payments': [],
        'expenses': [],
    }
    count = 0

    # ------------------ العقود ------------------
    if entity in ['all', 'contracts']:
        contracts_qs = Contract.objects.all().select_related('unit', 'tenant')
        if query:
            contracts_qs = contracts_qs.filter(
                Q(contract_number__icontains=query) |
                Q(tenant__name__icontains=query) |
                Q(unit__unit_number__icontains=query) |
                Q(notes__icontains=query)
            )
        if date_from:
            contracts_qs = contracts_qs.filter(start_date__gte=date_from)
        if date_to:
            contracts_qs = contracts_qs.filter(start_date__lte=date_to)
        if status == 'active':
            contracts_qs = contracts_qs.filter(is_active=True)
        elif status == 'inactive':
            contracts_qs = contracts_qs.filter(is_active=False)
        results['contracts'] = contracts_qs[:50]
        count += len(results['contracts'])

    # ------------------ الوحدات ------------------
    if entity in ['all', 'units']:
        units_qs = Unit.objects.all().select_related('sub_category')
        if query:
            units_qs = units_qs.filter(
                Q(unit_number__icontains=query) |
                Q(sub_category__name__icontains=query) |
                Q(notes__icontains=query)
            )
        if unit_type:
            units_qs = units_qs.filter(unit_type=unit_type)
        if status == 'rented':
            units_qs = units_qs.filter(is_rented=True)
        elif status == 'vacant':
            units_qs = units_qs.filter(is_rented=False)
        if main_category:
            units_qs = units_qs.filter(sub_category__main_category_id=main_category)
        if sub_category:
            units_qs = units_qs.filter(sub_category_id=sub_category)
        results['units'] = units_qs[:50]
        count += len(results['units'])

    # ------------------ المستأجرين (مع استبعاد المحذوفين) ------------------
    if entity in ['all', 'tenants']:
        # الأهم: نستبعد المحذوفين
        tenants_qs = Tenant.objects.filter(is_deleted=False)
        if query:
            tenants_qs = tenants_qs.filter(
                Q(name__icontains=query) |
                Q(identity_number__icontains=query) |
                Q(phone__icontains=query) |
                Q(email__icontains=query)
            )
        results['tenants'] = tenants_qs[:50]
        count += len(results['tenants'])

    # ------------------ الدفعات ------------------
    if entity in ['all', 'payments']:
        payments_qs = Payment.objects.all().select_related('contract__tenant', 'contract__unit')
        if query:
            payments_qs = payments_qs.filter(
                Q(contract__contract_number__icontains=query) |
                Q(contract__tenant__name__icontains=query) |
                Q(contract__unit__unit_number__icontains=query)
            )
        if date_from:
            payments_qs = payments_qs.filter(payment_date__gte=date_from)
        if date_to:
            payments_qs = payments_qs.filter(payment_date__lte=date_to)
        results['payments'] = payments_qs[:50]
        count += len(results['payments'])

    # ------------------ المصاريف ------------------
    if entity in ['all', 'expenses']:
        expenses_qs = Expense.objects.all().select_related('sub_category')
        if query:
            expenses_qs = expenses_qs.filter(
                Q(description__icontains=query) |
                Q(sub_category__name__icontains=query)
            )
        if date_from:
            expenses_qs = expenses_qs.filter(date__gte=date_from)
        if date_to:
            expenses_qs = expenses_qs.filter(date__lte=date_to)
        results['expenses'] = expenses_qs[:50]
        count += len(results['expenses'])

    # ------------------ بيانات القوائم المنسدلة ------------------
    main_categories = MainCategory.objects.all()
    sub_categories = SubCategory.objects.all()
    unit_types = Unit.UNIT_TYPES

    context = {
        'results': results,
        'count': count,
        'query': query,
        'entity': entity,
        'date_from': date_from,
        'date_to': date_to,
        'status': status,
        'unit_type': unit_type,
        'main_category': main_category,
        'sub_category': sub_category,
        'main_categories': main_categories,
        'sub_categories': sub_categories,
        'unit_types': unit_types,
    }
    return render(request, 'rentals/advanced_search.html', context)

@login_required
def export_expense_pdf(request):
    """تصدير تقرير المصاريف إلى PDF"""
    year = int(request.GET.get('year', datetime.now().year))
    month = int(request.GET.get('month', datetime.now().month))
    
    start_date = date(year, month, 1)
    if month == 12:
        end_date = date(year+1, 1, 1) - timedelta(days=1)
    else:
        end_date = date(year, month+1, 1) - timedelta(days=1)
    
    expenses = Expense.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).select_related('sub_category')
    
    total_expense = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    total_refundable = sum(e.tax_amount for e in expenses if e.tax_refundable)

    # ------------------- البحث عن خط عربي -------------------
    possible_paths = [
        '/Library/Fonts/Arial.ttf',
        '/System/Library/Fonts/Supplemental/Arial.ttf',
        'C:/Windows/Fonts/Arial.ttf',
        '/usr/share/fonts/truetype/msttcorefonts/Arial.ttf',
    ]
    font_registered = False
    for path in possible_paths:
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont('Arabic', path))
            font_registered = True
            break
    if not font_registered:
        pdfmetrics.registerFont(TTFont('Arabic', 'Helvetica'))
    # ---------------------------------------------------------

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="expense_report_{year}_{month}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=30, bottomMargin=20)
    elements = []
    
    styles = getSampleStyleSheet()
    styles['Title'].fontName = 'Arabic'
    title_style = styles['Title']
    title_style.alignment = 1
    
    # عنوان التقرير (مع reshape)
    title_text = f"تقرير المصاريف - {get_month_name(month)} {year}"
    title = Paragraph(prepare_arabic_text(title_text), title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.5*cm))
    
    # معلومات الفترة والإجمالي (مع reshape)
    info_style = ParagraphStyle('info', parent=styles['Normal'], fontName='Arabic', alignment=1)
    info_text = f"الفترة: {start_date} إلى {end_date} | إجمالي المصاريف: {total_expense:,.2f} | ضريبة مستردة: {total_refundable:,.2f}"
    info = Paragraph(prepare_arabic_text(info_text), info_style)
    elements.append(info)
    elements.append(Spacer(1, 0.5*cm))
    
    # بيانات الجدول
    data = []
    headers = ['التاريخ', 'القسم', 'البيان', 'المبلغ', 'ضريبة مستردة']
    reshaped_headers = [prepare_arabic_text(h) for h in headers]
    data.append(reshaped_headers)
    
    for e in expenses:
        row = [
            str(e.date),
            prepare_arabic_text(e.sub_category.name),
            prepare_arabic_text(e.description),
            f"{e.amount:,.2f}",
            f"{e.tax_amount:,.2f}" if e.tax_refundable else prepare_arabic_text("-"),
        ]
        data.append(row)
    
    table = Table(data, colWidths=[3*cm, 4*cm, 5*cm, 3*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Arabic'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response


